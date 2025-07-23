import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from utils import count_unique_products_per_sheet, load_all_sheets, remove_flavor_variants, find_similar_products, \
    find_normal_cases, pairs_to_unique_products, process_excel_for_duplicates, find_internal_duplicates

if __name__=="__main__":
    # all data
    excel_path = "/home/viktoria/Downloads/ARBOLES ALIMENTOS EJERCICIO KABELLI.xlsx"
    df_all = load_all_sheets(excel_path)
    df_all = df_all.drop_duplicates(["Marca", "Nombre SKU", "SKU", "Sheet"])

    # exact and partial matches
    correct_products = find_similar_products(df_all, 88, different_sku=False)
    correct_products = correct_products.copy()
    correct_products = remove_flavor_variants(correct_products)
    columns_to_show = [col for col in correct_products.columns if col not in ['Numbers 1', 'Numbers 2']]
    correct_products = correct_products.loc[:, columns_to_show]
    exact_matches = correct_products[correct_products['Similarity'] == 100]
    partial_matches = correct_products[correct_products['Similarity'] < 100]
    exact_matches_counts = count_unique_products_per_sheet(exact_matches)
    partial_matches_counts = count_unique_products_per_sheet(partial_matches)

    # same product with different sku
    confident, needs_review = process_excel_for_duplicates(
        excel_path,
        confidence_threshold=93,
        low_confidence_threshold=85
    )
    condident_prod = pairs_to_unique_products(confident)
    need_review_prod = pairs_to_unique_products(needs_review)
    all_different_sku = pd.concat([condident_prod, need_review_prod])
    all_different_sku = all_different_sku.drop_duplicates(["Nombre SKU", "Marca", "SKU"])

    unique_sheets = all_different_sku['Sheet'].unique()
    unique_sheets = [sheet for sheet in unique_sheets if sheet != "Prilogic Arbol_24_25"]

    all_different_sku_count = {}

    for sheet in unique_sheets:
        sub_df = all_different_sku[all_different_sku['Sheet'] == sheet]
        all_different_sku_count[sheet] = sub_df.shape[0]

    # duplicates
    duplicates = find_internal_duplicates(df_all)
    unique_sheets = duplicates['Sheet'].unique()
    unique_sheets = [sheet for sheet in unique_sheets if sheet != "Prilogic Arbol_24_25"]

    duplicates_counts = {}

    for sheet in unique_sheets:
        sub_df = duplicates[duplicates['Sheet'] == sheet]
        duplicates_counts[sheet] = sub_df.shape[0]

    filtered = find_normal_cases(excel_path)
    filtered = filtered.rename(columns={"Sheet": "Subempresa"})
    filtered = filtered.drop_duplicates()
    filtered_count = filtered.shape[0]

    filtered = find_normal_cases(excel_path)
    filtered = filtered.rename(columns={"Sheet": "Subempresa"})
    filtered = filtered.drop_duplicates()
    filtered_count = filtered.shape[0]

    unique_sheets = filtered['Subempresa'].unique()
    unique_sheets = [sheet for sheet in unique_sheets if sheet != "Prilogic Arbol_24_25"]

    filtered_counts = {}

    for sheet in unique_sheets:
        sub_df = filtered[filtered['Subempresa'] == sheet]
        filtered_counts[sheet] = sub_df.shape[0]

    dictionaries = [exact_matches_counts, partial_matches_counts, duplicates_counts, all_different_sku_count,
                    filtered_counts]

    results = dict()
    for dct in dictionaries:
        for key in dct.keys():
            if not key in results.keys():
                results[key] = dct[key]
            else:
                results[key] += dct[key]

    dictionaries = [
        duplicates_counts,
        exact_matches_counts,
        partial_matches_counts,
        all_different_sku_count,
        filtered_counts,
        results
    ]

    column_names = [
        "Duplicates",
        "Same SKU, same name",
        "Same SKU, similar name",
        "Same product, different SKU",
        "Unique Products",
        "Total"
    ]

    df = pd.DataFrame(dictionaries, index=column_names).T.fillna(0).astype(int)
    df.index.name = "Company"
    print(results)


    df_reset = df.reset_index()

    excel_path = "Products for each company.xlsx"
    df_reset.to_excel(excel_path, index=False)


    column_colors = {
        "Company": "D3D3D3",                # Light Gray
        "Total": "D3D3D3",                  # Light Gray
        "Duplicates": "A52A2A",            # Brown
        "Same SKU, same name": "87CEEB",   # Blue
        "Same SKU, similar name": "9370DB",# Purple
        "Same product, different SKU": "90EE90", # Green
        "Unique Products": "FF7F7F"        # Red
    }

    column_widths = {
        "Company": 30,
        "Duplicates": 15,
        "Same SKU, same name": 22,
        "Same SKU, similar name": 25,
        "Same product, different SKU": 28,
        "Unique Products": 18,
        "Total": 10
    }

    wb = load_workbook(excel_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    for col_idx, col_name in enumerate(header, start=1):

        color = column_colors.get(col_name)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = fill

        width = column_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "B2"
    # Save
    wb.save(excel_path)
