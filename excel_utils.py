import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def prepare_df(df, key):
    """Helper function for save_products_pairs_to_excel"""
    df = df.copy()
    CATEGORY_MAP = {
        "exact_matches": "same name, same sku",
        "partial_matches": "similar name, same sku",
        "confident": "similar name, different sku",
        "needs_review": "similar name, different sku",
        "filtered": "unique product"
    }
    if key == "filtered":
        df = df.rename(columns={
            "Nombre SKU": "Nombre SKU 1",
            "SKU": "SKU 1",
            "Subempresa": "Subempresa 1"
        })

        df["Nombre SKU 2"] = None
        df["SKU 2"] = None
        df["Subempresa 2"] = None
        df["Similarity"] = None
    else:
        df = df.rename(columns={
            "Sheet 1": "Subempresa 1",
            "Sheet 2": "Subempresa 2"
        })

    df["category"] = CATEGORY_MAP[key]

    desired_columns = [
        "Marca", "Nombre SKU 1", "SKU 1", "Subempresa 1",
        "Nombre SKU 2", "SKU 2", "Subempresa 2", "Similarity", "category"
    ]
    df = df.loc[:, desired_columns]

    return df


def save_products_pairs_to_excel(exact_matches, partial_matches, confident, needs_review, filtered):
    """
    Save matched product pairs to an Excel file with coloring and consistent formatting.
    """

    COLOR_MAP = {
        "exact_matches": "BDD7EE",  # light blue
        "partial_matches": "D9D2E9",  # light purple
        "confident": "C6E0B4",  # light green
        "needs_review": "C6E0B4",  # light green (same as confident)
        "filtered": "F4CCCC",  # light red
    }

    tables = {
        "exact_matches": exact_matches,
        "partial_matches": partial_matches,
        "confident": confident,
        "needs_review": needs_review,
        "filtered": filtered,
    }

    dfs = []
    order = ["exact_matches", "partial_matches", "confident", "needs_review", "filtered"]
    for key in order:
        df = prepare_df(tables[key], key)
        dfs.append(df)

    combined_df = pd.concat(dfs, ignore_index=True)

    output_file = "/home/viktoria/Downloads/pairs_of_products.xlsx"
    combined_df.to_excel(output_file, index=False, sheet_name="Subempresa")

    wb = load_workbook(output_file)
    ws = wb["Subempresa"]

    row_to_color = {}
    row_idx = 2

    for key, df in zip(order, dfs):
        for _ in range(len(df)):
            row_to_color[row_idx] = COLOR_MAP[key]
            row_idx += 1

    for row, color in row_to_color.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 4
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)
    print(f"Excel file saved as '{output_file}'")
